import io

import pytest
from openpyxl import Workbook

from apps.datasets.exceptions import EmptyFileError, InvalidExcelError, NoHeaderError
from apps.datasets.models import DatasetColumn
from apps.datasets.services.ingestion import PREVIEW_SAMPLE_SIZE, preview_file


def csv_file(content: str) -> io.BytesIO:
    return io.BytesIO(content.encode("utf-8"))


class TestPreviewFile:
    def test_returns_detected_schema_and_sample_rows(self):
        content = csv_file(
            "name,email,campus\nSarah,sarah@example.com,Paris\nLea,lea@example.com,Lyon\n"
        )

        result = preview_file(content, "leads.csv")

        assert result["row_count"] == 2
        assert [c["name_normalized"] for c in result["columns"]] == ["name", "email", "campus"]
        assert result["columns"][1]["detected_type"] == DatasetColumn.ColumnType.EMAIL
        assert result["sample_rows"][0] == {
            "name": "Sarah",
            "email": "sarah@example.com",
            "campus": "Paris",
        }

    def test_does_not_persist_anything(self, db):
        from apps.datasets.models import Dataset, DatasetRow

        preview_file(csv_file("name\nSarah\n"), "leads.csv")

        assert Dataset.objects.count() == 0
        assert DatasetColumn.objects.count() == 0
        assert DatasetRow.objects.count() == 0

    def test_caps_the_sample_at_preview_sample_size(self):
        rows = "\n".join(f"row{i}" for i in range(PREVIEW_SAMPLE_SIZE + 5))
        content = csv_file(f"name\n{rows}\n")

        result = preview_file(content, "leads.csv")

        assert result["row_count"] == PREVIEW_SAMPLE_SIZE + 5
        assert len(result["sample_rows"]) == PREVIEW_SAMPLE_SIZE

    def test_dispatches_to_xlsx_parsing_by_extension(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["name", "email"])
        sheet.append(["Sarah", "sarah@example.com"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        result = preview_file(buffer, "leads.xlsx")

        assert result["row_count"] == 1
        assert [c["name_normalized"] for c in result["columns"]] == ["name", "email"]

    def test_csv_reports_no_available_sheets(self):
        result = preview_file(csv_file("name\nSarah\n"), "leads.csv")
        assert result["available_sheets"] == []

    def test_xlsx_lists_available_sheets_and_previews_the_first_by_default(self):
        workbook = Workbook()
        workbook.remove(workbook.active)
        students = workbook.create_sheet("Students")
        students.append(["name"])
        students.append(["Sarah"])
        staff = workbook.create_sheet("Staff")
        staff.append(["name", "role"])
        staff.append(["Marc", "Teacher"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        result = preview_file(buffer, "school.xlsx")

        assert result["available_sheets"] == ["Students", "Staff"]
        assert [c["name_normalized"] for c in result["columns"]] == ["name"]

    def test_xlsx_can_preview_a_specific_sheet(self):
        workbook = Workbook()
        workbook.remove(workbook.active)
        students = workbook.create_sheet("Students")
        students.append(["name"])
        staff = workbook.create_sheet("Staff")
        staff.append(["name", "role"])
        staff.append(["Marc", "Teacher"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        result = preview_file(buffer, "school.xlsx", sheet_name="Staff")

        assert [c["name_normalized"] for c in result["columns"]] == ["name", "role"]
        assert result["row_count"] == 1

    def test_raises_on_empty_file(self):
        with pytest.raises(EmptyFileError):
            preview_file(csv_file(""), "empty.csv")

    def test_raises_on_missing_header(self):
        with pytest.raises(NoHeaderError):
            preview_file(csv_file("\ndata\n"), "bad.csv")

    def test_raises_on_corrupt_xlsx(self):
        with pytest.raises(InvalidExcelError):
            preview_file(io.BytesIO(b"not an xlsx"), "bad.xlsx")
