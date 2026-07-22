import datetime
import io
import zipfile

import pytest
from django.contrib.auth import get_user_model
from openpyxl import Workbook

from apps.datasets.exceptions import EmptyFileError, InvalidExcelError, NoHeaderError
from apps.datasets.models import Dataset, DatasetColumn, DatasetRow
from apps.datasets.services.ingestion import ingest_xlsx_file, list_workbook_sheets

pytestmark = pytest.mark.django_db

User = get_user_model()


def xlsx_file(rows: list[list]) -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def xlsx_file_multi_sheet(sheets: dict[str, list[list]]) -> io.BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets.items():
        sheet = workbook.create_sheet(sheet_name)
        for row in rows:
            sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@pytest.fixture
def user():
    return User.objects.create_user(username="modeste", password="pass1234")


@pytest.fixture
def dataset(user):
    return Dataset.objects.create(
        owner=user, name="inscriptions", original_filename="inscriptions.xlsx"
    )


class TestIngestXlsxFile:
    def test_creates_columns_and_rows_from_the_first_sheet(self, dataset):
        content = xlsx_file(
            [
                ["name", "email", "campus"],
                ["Sarah", "sarah@example.com", "Paris"],
                ["Lea", "lea@example.com", "Lyon"],
            ]
        )
        ingest_xlsx_file(dataset, content)

        dataset.refresh_from_db()
        assert dataset.status == Dataset.Status.READY
        assert dataset.row_count == 2
        assert dataset.column_count == 3

        columns = list(DatasetColumn.objects.filter(dataset=dataset).order_by("position"))
        assert [c.name_normalized for c in columns] == ["name", "email", "campus"]
        assert columns[1].detected_type == DatasetColumn.ColumnType.EMAIL

        rows = list(DatasetRow.objects.filter(dataset=dataset).order_by("row_index"))
        assert rows[0].data == {"name": "Sarah", "email": "sarah@example.com", "campus": "Paris"}

    def test_normalizes_headers_same_as_csv(self, dataset):
        content = xlsx_file([["Full Name", "Email", "email"], ["Sarah", "a@b.com", "c@d.com"]])
        ingest_xlsx_file(dataset, content)

        columns = list(DatasetColumn.objects.filter(dataset=dataset).order_by("position"))
        assert [c.name_normalized for c in columns] == ["full_name", "email", "email_2"]

    def test_numeric_cells_are_stringified_without_trailing_zero(self, dataset):
        content = xlsx_file([["name", "age"], ["Sarah", 42], ["Lea", 17.5]])
        ingest_xlsx_file(dataset, content)

        rows = list(DatasetRow.objects.filter(dataset=dataset).order_by("row_index"))
        assert rows[0].data["age"] == "42"
        assert rows[1].data["age"] == "17.5"

        age_column = DatasetColumn.objects.get(dataset=dataset, name_normalized="age")
        assert age_column.detected_type == DatasetColumn.ColumnType.NUMBER

    def test_date_cells_are_stringified_as_iso_dates(self, dataset):
        content = xlsx_file([["name", "joined"], ["Sarah", datetime.date(2027, 1, 15)]])
        ingest_xlsx_file(dataset, content)

        row = DatasetRow.objects.get(dataset=dataset)
        assert row.data["joined"] == "2027-01-15"
        joined_column = DatasetColumn.objects.get(dataset=dataset, name_normalized="joined")
        assert joined_column.detected_type == DatasetColumn.ColumnType.DATE

    def test_boolean_cells_are_stringified_as_true_false(self, dataset):
        content = xlsx_file([["name", "active"], ["Sarah", True], ["Lea", False]])
        ingest_xlsx_file(dataset, content)

        rows = list(DatasetRow.objects.filter(dataset=dataset).order_by("row_index"))
        assert rows[0].data["active"] == "true"
        assert rows[1].data["active"] == "false"

    def test_empty_cells_become_empty_strings(self, dataset):
        content = xlsx_file([["name", "campus"], ["Sarah", None], ["Lea", "Lyon"]])
        ingest_xlsx_file(dataset, content)

        rows = list(DatasetRow.objects.filter(dataset=dataset).order_by("row_index"))
        assert rows[0].data["campus"] == ""

    def test_raises_on_empty_file(self, dataset):
        with pytest.raises(EmptyFileError):
            ingest_xlsx_file(dataset, io.BytesIO(b""))

    def test_raises_when_first_sheet_has_no_rows_at_all(self, dataset):
        buffer = io.BytesIO()
        Workbook().save(buffer)
        buffer.seek(0)

        with pytest.raises(NoHeaderError):
            ingest_xlsx_file(dataset, buffer)
        dataset.refresh_from_db()
        assert dataset.status == Dataset.Status.FAILED

    def test_raises_on_corrupt_or_non_xlsx_content(self, dataset):
        with pytest.raises(InvalidExcelError):
            ingest_xlsx_file(dataset, io.BytesIO(b"this is not an xlsx file"))
        dataset.refresh_from_db()
        assert dataset.status == Dataset.Status.FAILED
        assert dataset.failure_reason

    def test_raises_on_a_valid_zip_that_isnt_a_workbook(self, dataset):
        # A well-formed zip archive, but missing the xl/workbook.xml (etc.) parts
        # openpyxl requires — this is a different failure mode than a non-zip file
        # (BadZipFile) or a workbook with no sheets, and must be caught just the same.
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w") as archive:
            archive.writestr("not_a_workbook.txt", "hello")
        buffer.seek(0)

        with pytest.raises(InvalidExcelError):
            ingest_xlsx_file(dataset, buffer)
        dataset.refresh_from_db()
        assert dataset.status == Dataset.Status.FAILED

    def test_raises_on_malformed_worksheet_xml(self, dataset):
        # A workbook whose worksheet XML is truncated mid-stream — openpyxl's read_only
        # mode parses sheet XML lazily while iterating rows, so this must be caught
        # around the row loop too, not just around load_workbook() itself.
        source = xlsx_file([["name", "email"], ["Sarah", "sarah@example.com"]])
        with zipfile.ZipFile(source) as archive:
            names = archive.namelist()
            sheet_name = next(n for n in names if n.startswith("xl/worksheets/sheet"))
            parts = {n: archive.read(n) for n in names}

        parts[sheet_name] = parts[sheet_name][: len(parts[sheet_name]) // 2]

        corrupted = io.BytesIO()
        with zipfile.ZipFile(corrupted, "w") as archive:
            for name, content in parts.items():
                archive.writestr(name, content)
        corrupted.seek(0)

        with pytest.raises(InvalidExcelError):
            ingest_xlsx_file(dataset, corrupted)
        dataset.refresh_from_db()
        assert dataset.status == Dataset.Status.FAILED

    def test_raises_when_workbook_has_no_worksheets(self, dataset, monkeypatch):
        content = xlsx_file([["name"], ["Sarah"]])

        from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook

        monkeypatch.setattr(OpenpyxlWorkbook, "worksheets", property(lambda self: []))

        with pytest.raises(InvalidExcelError):
            ingest_xlsx_file(dataset, content)
        dataset.refresh_from_db()
        assert dataset.status == Dataset.Status.FAILED

    def test_raises_when_row_count_exceeds_the_configured_limit(self, dataset, settings):
        settings.FILEBRIDGE_MAX_XLSX_ROWS = 2
        content = xlsx_file([["name"], ["a"], ["b"], ["c"]])  # header + 3 data rows > 2

        with pytest.raises(InvalidExcelError):
            ingest_xlsx_file(dataset, content)
        dataset.refresh_from_db()
        assert dataset.status == Dataset.Status.FAILED
        assert "rows" in dataset.failure_reason.lower()

    def test_ingests_a_specific_named_sheet(self, dataset):
        content = xlsx_file_multi_sheet(
            {
                "Students": [["name"], ["Sarah"]],
                "Staff": [["name", "role"], ["Marc", "Teacher"], ["Lea", "Admin"]],
            }
        )
        ingest_xlsx_file(dataset, content, sheet_name="Staff")

        dataset.refresh_from_db()
        assert dataset.status == Dataset.Status.READY
        assert dataset.row_count == 2
        assert dataset.column_count == 2

    def test_defaults_to_the_first_sheet_when_none_is_given(self, dataset):
        content = xlsx_file_multi_sheet(
            {
                "Students": [["name"], ["Sarah"]],
                "Staff": [["name", "role"], ["Marc", "Teacher"]],
            }
        )
        ingest_xlsx_file(dataset, content)

        dataset.refresh_from_db()
        assert dataset.column_count == 1  # "Students" sheet, the first one

    def test_raises_a_clear_error_for_an_unknown_sheet_name(self, dataset):
        content = xlsx_file_multi_sheet({"Students": [["name"], ["Sarah"]]})

        with pytest.raises(InvalidExcelError, match="Nonexistent"):
            ingest_xlsx_file(dataset, content, sheet_name="Nonexistent")
        dataset.refresh_from_db()
        assert dataset.status == Dataset.Status.FAILED


class TestListWorkbookSheets:
    def test_returns_sheet_names_in_workbook_order(self):
        content = xlsx_file_multi_sheet(
            {"Students": [["name"]], "Staff": [["name"]], "Alumni": [["name"]]}
        )
        assert list_workbook_sheets(content.read()) == ["Students", "Staff", "Alumni"]

    def test_raises_on_corrupt_content(self):
        with pytest.raises(InvalidExcelError):
            list_workbook_sheets(b"not an xlsx file")
