"""Ingestion: parse an uploaded CSV or Excel file into Dataset/DatasetColumn/DatasetRow rows.

Both formats are normalized to the same shape (a list of header strings, a list of
string rows) before sharing the exact same column-detection and row-creation logic —
see _parse_csv/_parse_xlsx vs _ingest_rows below.
"""

import csv
import io
import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import date, datetime, time, timedelta

from django.conf import settings
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from apps.datasets.exceptions import (
    EmptyFileError,
    InvalidCsvError,
    InvalidExcelError,
    NoHeaderError,
)
from apps.datasets.models import Dataset, DatasetColumn, DatasetRow
from apps.datasets.services.type_detection import detect_column_type

_NON_ALNUM_RE = re.compile(r"[^a-z0-9]+")


def normalize_column_name(raw: str, seen: dict | None = None) -> str:
    """Turn a raw header into a slug-safe, snake_case column name.

    When `seen` (a dict shared across the whole header) is provided, repeated
    names are disambiguated as name, name_2, name_3, ...
    """
    slug = _NON_ALNUM_RE.sub("_", raw.strip().lower()).strip("_") or "column"

    if seen is None:
        return slug

    if slug not in seen:
        seen[slug] = 1
        return slug

    seen[slug] += 1
    candidate = f"{slug}_{seen[slug]}"
    while candidate in seen:
        seen[slug] += 1
        candidate = f"{slug}_{seen[slug]}"
    seen[candidate] = 1
    return candidate


def ingest_csv_file(dataset: Dataset, file_obj) -> None:
    """Parse `file_obj` as CSV and populate columns/rows for `dataset`.

    Marks the dataset FAILED (with a reason) and re-raises on unusable input,
    so callers can surface a clean 4xx instead of a 500.
    """
    raw_bytes = _read_or_fail_empty(dataset, file_obj)

    try:
        text = raw_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        _fail(dataset, "The uploaded file isn't valid UTF-8 text.")
        raise InvalidCsvError("The uploaded file isn't valid UTF-8 text.") from exc

    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader, [])
    except csv.Error as exc:
        _fail(dataset, "The uploaded file isn't valid CSV.")
        raise InvalidCsvError("The uploaded file isn't valid CSV.") from exc

    if not header or all(not cell.strip() for cell in header):
        _fail(dataset, "The uploaded file has no header row.")
        raise NoHeaderError("The uploaded file has no header row.")

    try:
        data_rows = list(reader)
    except csv.Error as exc:
        _fail(dataset, "The uploaded file isn't valid CSV.")
        raise InvalidCsvError("The uploaded file isn't valid CSV.") from exc

    _ingest_rows(dataset, header, data_rows)


_INVALID_WORKBOOK_ERRORS = (
    InvalidFileException,
    KeyError,
    IndexError,
    OSError,
    zipfile.BadZipFile,
    ET.ParseError,
)


def ingest_xlsx_file(dataset: Dataset, file_obj) -> None:
    """Parse `file_obj` as an Excel workbook (first sheet only) and populate
    columns/rows for `dataset`. Same failure-handling contract as ingest_csv_file.
    """
    raw_bytes = _read_or_fail_empty(dataset, file_obj)

    workbook = None
    try:
        workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        sheet = workbook.worksheets[0]

        # openpyxl's read_only mode parses each sheet's XML lazily while iterating, so a
        # malformed worksheet part can raise here too, not just from load_workbook() —
        # and a hard row cap protects against a small file that decompresses into a
        # huge number of rows (a zip is, after all, a compressed archive).
        sheet_rows = []
        for row in sheet.iter_rows(values_only=True):
            if len(sheet_rows) >= settings.FILEBRIDGE_MAX_XLSX_ROWS:
                raise InvalidExcelError(
                    f"The workbook has more than {settings.FILEBRIDGE_MAX_XLSX_ROWS} rows."
                )
            sheet_rows.append([_stringify_cell(cell) for cell in row])
    except InvalidExcelError as exc:
        _fail(dataset, str(exc))
        raise
    except _INVALID_WORKBOOK_ERRORS as exc:
        _fail(dataset, "The uploaded file isn't a valid Excel (.xlsx) workbook.")
        raise InvalidExcelError("The uploaded file isn't a valid Excel (.xlsx) workbook.") from exc
    finally:
        if workbook is not None:
            workbook.close()

    header = sheet_rows[0] if sheet_rows else []
    if not header or all(not cell.strip() for cell in header):
        _fail(dataset, "The uploaded file has no header row.")
        raise NoHeaderError("The uploaded file has no header row.")

    _ingest_rows(dataset, header, sheet_rows[1:])


def _stringify_cell(value) -> str:
    """Normalize an openpyxl cell value to the same plain-string shape a CSV cell has."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value.total_seconds())
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _read_or_fail_empty(dataset: Dataset, file_obj) -> bytes:
    raw_bytes = file_obj.read()
    if not raw_bytes or not raw_bytes.strip():
        _fail(dataset, "The uploaded file is empty.")
        raise EmptyFileError("The uploaded file is empty.")
    return raw_bytes


def _ingest_rows(dataset: Dataset, header: list[str], data_rows: list[list[str]]) -> None:
    """Shared second half of ingestion: normalize headers, detect types, bulk-create."""
    seen: dict[str, int] = {}
    normalized_names = [normalize_column_name(cell, seen=seen) for cell in header]

    columns_values: list[list[str]] = [[] for _ in normalized_names]
    for row in data_rows:
        for i in range(len(normalized_names)):
            columns_values[i].append(row[i] if i < len(row) else "")
    detected_types = [detect_column_type(values) for values in columns_values]

    with transaction.atomic():
        DatasetColumn.objects.filter(dataset=dataset).delete()
        DatasetRow.objects.filter(dataset=dataset).delete()

        columns = DatasetColumn.objects.bulk_create(
            DatasetColumn(
                dataset=dataset,
                name_original=original,
                name_normalized=normalized,
                detected_type=col_type,
                position=position,
            )
            for position, (original, normalized, col_type) in enumerate(
                zip(header, normalized_names, detected_types)
            )
        )

        rows = DatasetRow.objects.bulk_create(
            DatasetRow(
                dataset=dataset,
                row_index=row_index,
                data={
                    normalized_names[i]: (row[i] if i < len(row) else "")
                    for i in range(len(normalized_names))
                },
            )
            for row_index, row in enumerate(data_rows)
        )

        dataset.status = Dataset.Status.READY
        dataset.row_count = len(rows)
        dataset.column_count = len(columns)
        dataset.failure_reason = ""
        dataset.save(update_fields=["status", "row_count", "column_count", "failure_reason"])


def _fail(dataset: Dataset, reason: str) -> None:
    dataset.status = Dataset.Status.FAILED
    dataset.failure_reason = reason
    dataset.save(update_fields=["status", "failure_reason"])
